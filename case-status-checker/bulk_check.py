"""
ตรวจสอบสถานะ "บุคคลล้มละลาย" จากเว็บไซต์กรมบังคับคดี (led.go.th) แบบกลุ่ม (bulk)
โดยอ่านรายชื่อ/เลขบัตรประชาชนจากไฟล์ Excel แล้วกรอกฟอร์มค้นหาให้อัตโนมัติทีละคน

เนื่องจากฟอร์มค้นหาของกรมบังคับคดีมี CAPTCHA ป้องกันการยิงคำขอซ้ำๆ อัตโนมัติ
สคริปต์นี้จึง "ไม่พยายามถอดรหัส CAPTCHA เอง" แต่จะเปิดเบราว์เซอร์จริงให้ผู้ใช้เห็น
และหยุดรอให้ผู้ใช้พิมพ์รหัสที่เห็นในภาพทีละรอบ ส่วนที่เหลือ (ล็อกอิน, กรอกฟอร์ม,
กดค้นหา, อ่านผลลัพธ์, บันทึกไฟล์) ทำให้อัตโนมัติทั้งหมด

วิธีใช้งาน:
    pip install -r requirements.txt
    playwright install chromium
    python bulk_check.py --input CardID.xlsx --sheet 0769 --output results.xlsx

ระหว่างรัน:
    - สคริปต์จะเปิดเบราว์เซอร์เปล่าขึ้นมา ให้ผู้ใช้นำทาง/ล็อกอินเองตามปกติ
      (พิมพ์ URL, ล็อกอิน, คลิกเมนู) จนกว่าจะเปิดหน้า "สอบถามบุคคลล้มละลาย"
      (WEB3Q010) ได้สำเร็จ 1 ครั้ง แล้วกด Enter ในหน้าต่าง terminal เพื่อให้
      สคริปต์ทำงานต่อ
    - จากนั้นสำหรับแต่ละรายชื่อ สคริปต์จะกรอกเลขบัตรประชาชนให้ แล้วหยุดรอให้พิมพ์
      รหัส CAPTCHA ที่เห็นในเบราว์เซอร์ (หรือดูจากไฟล์ captcha.png ที่บันทึกไว้)
    - ผลลัพธ์จะถูกบันทึกลงไฟล์ Excel ทันทีหลังตรวจแต่ละราย ถ้าสคริปต์ถูกปิดกลางคัน
      รันคำสั่งเดิมซ้ำได้เลย มันจะข้ามรายชื่อที่ตรวจสอบไปแล้วในไฟล์ผลลัพธ์โดยอัตโนมัติ

หมายเหตุสำคัญ: สคริปต์นี้เขียนจาก HTML จริงของหน้าฟอร์ม "สอบถามบุคคลล้มละลาย"
(WEB3Q010) แต่ยังไม่เคยรันกับระบบจริง แนะนำให้ทดสอบกับรายชื่อ 3-5 รายก่อน
(--limit 5) แล้วตรวจสอบผลลัพธ์ให้ตรงกับที่เห็นในเบราว์เซอร์จริง ก่อนรันเต็มจำนวน
"""

import argparse
import datetime as dt
import sys
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

FORM_URL = "https://ledwebsite.led.go.th/ledweb/led/web/system/WEB3Q010Action.do"

STATUS_CLEAR = "ไม่พบคดี"
STATUS_FOUND = "พบคดี"
STATUS_ERROR = "ตรวจสอบไม่สำเร็จ"

OUTPUT_HEADERS = [
    "เลขทะเบียนสมาชิก",
    "เลขบัตรประชาชน",
    "ชื่อ-นามสกุล (จากไฟล์)",
    "สถานะคดี",
    "รายละเอียดคดีที่พบ",
    "ตรวจสอบเมื่อ",
]

MAX_CAPTCHA_ATTEMPTS = 5


def read_input_rows(path, sheet_name, id_col, member_col, name_col):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        raw_id = row[id_col]
        if raw_id is None:
            continue
        id_str = str(raw_id).strip()
        if not id_str.isdigit() or len(id_str) != 13:
            continue
        member_no = str(row[member_col]).strip() if row[member_col] is not None else ""
        name = str(row[name_col]).strip() if row[name_col] is not None else ""
        rows.append({"id": id_str, "member_no": member_no, "name": name})
    return rows


def load_done_ids(output_path):
    done = set()
    if not Path(output_path).exists():
        return done
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[1]:
            done.add(str(row[1]).strip())
    return done


def open_or_create_output(output_path):
    if Path(output_path).exists():
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(OUTPUT_HEADERS)
    return wb, ws


def append_result(wb, ws, output_path, record, status, detail):
    ws.append([
        record["member_no"],
        record["id"],
        record["name"],
        status,
        detail,
        dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])
    wb.save(output_path)


def setup_dialog_capture(page):
    """ดักจับ popup alert() จริงๆ ของเบราว์เซอร์ (ไม่ใช่การเดาจากข้อความในหน้า HTML
    เพราะข้อความแจ้งเตือนบางอย่าง เช่น 'กรุณาระบุ...' ถูกฝังอยู่ใน <script> ของทุกหน้า
    เสมออยู่แล้ว ต่อให้ไม่มี error จริงก็จะเจอข้อความนั้นใน page.content())."""
    state = {"message": None}

    def handler(dialog):
        state["message"] = dialog.message
        dialog.accept()

    page.on("dialog", handler)
    return state


def fill_search_form(page, record):
    page.fill("input[name='dfId']", record["id"])
    page.check("input[name='typeQuery'][value='1']")  # ตรงตัว (exact match) สำหรับค้นด้วยเลขบัตร


def screenshot_captcha(page, path):
    try:
        page.locator("#imgCaptcha").screenshot(path=path)
        return True
    except Exception:
        return False


def refresh_captcha(page):
    try:
        page.click("input[onclick*='refresh']")
        page.wait_for_timeout(500)
    except Exception:
        pass


def submit_query(page):
    with page.expect_navigation(wait_until="networkidle", timeout=30000):
        page.click("#Image31")


NOT_FOUND_TEXT = "ไม่พบรายการที่ค้นหา"


def parse_results_table(page):
    """คืนค่า (status, detail_text) จากตาราง lkPccBankruptTable หลังค้นหา."""
    table_text = page.locator("#lkPccBankruptTable").inner_text()

    if NOT_FOUND_TEXT in table_text:
        return STATUS_CLEAR, ""

    rows = page.locator(
        "#scrollContent_lkPccBankruptTable tr:not(.tableEmptyCell)"
    )
    count = rows.count()
    cases = []
    for i in range(count):
        cells = rows.nth(i).locator("td")
        if cells.count() < 6:
            continue
        full_name = cells.nth(1).inner_text().strip()
        if not full_name or full_name == "\xa0":
            continue
        recv_case = cells.nth(2).inner_text().strip()
        court = cells.nth(3).inner_text().strip()
        black_no = cells.nth(4).inner_text().strip()
        red_no = cells.nth(5).inner_text().strip()
        cases.append(
            f"{full_name} | เรื่องที่ {recv_case} | {court} | ดำที่ {black_no} | แดงที่ {red_no}"
        )

    if cases:
        return STATUS_FOUND, "; ".join(cases)

    # ไม่เจอทั้งข้อความ "ไม่พบรายการที่ค้นหา" และไม่เจอแถวข้อมูลจริง — ไม่ควรเดาว่า
    # "ไม่พบคดี" เฉยๆ เพราะอาจเป็นกรณี CAPTCHA ผิดที่ไม่มี alert เด้ง (ยังไม่ยืนยัน)
    # ให้รายงานเป็นสถานะที่ต้องตรวจสอบมือแทน เพื่อความปลอดภัย
    return STATUS_ERROR, "ไม่พบข้อความยืนยันผลลัพธ์ที่คาดไว้ ควรตรวจสอบด้วยตนเอง"


def process_one(page, record, captcha_path, dialog_state):
    """คืนค่า (status, detail)"""
    for attempt in range(1, MAX_CAPTCHA_ATTEMPTS + 1):
        page.goto(FORM_URL, wait_until="networkidle")
        fill_search_form(page, record)
        screenshot_captcha(page, captcha_path)

        print(
            f"\n  ผู้ตรวจสอบ: {record['name']}  เลขบัตร: {record['id']}"
            f"  (ครั้งที่ลอง {attempt}/{MAX_CAPTCHA_ATTEMPTS})"
        )
        print(f"  ดูรหัส CAPTCHA ในเบราว์เซอร์ หรือเปิดไฟล์: {captcha_path}")
        code = input("  พิมพ์รหัส CAPTCHA (พิมพ์ 'r' เพื่อขอรหัสใหม่, 's' เพื่อข้ามรายนี้): ").strip()

        if code.lower() == "s":
            return STATUS_ERROR, "ผู้ใช้เลือกข้ามรายนี้"
        if code.lower() == "r":
            continue  # goto ใหม่รอบหน้าจะได้ captcha ใหม่อยู่แล้ว

        page.fill("input[name='imgCode']", code)
        actual = page.input_value("input[name='imgCode']")
        if actual != code:
            print(f"  [เตือน] ช่องรหัสยืนยันอ่านได้ค่า '{actual}' ไม่ตรงกับที่พิมพ์ '{code}' กำลังลองกรอกใหม่")
            page.fill("input[name='imgCode']", code)
            actual = page.input_value("input[name='imgCode']")
            print(f"  [ตรวจสอบ] ค่าในช่องหลังกรอกใหม่: '{actual}'")

        dialog_state["message"] = None
        try:
            submit_query(page)
        except Exception as e:
            if dialog_state["message"]:
                print(f"  [ไม่ผ่าน] เบราว์เซอร์แจ้งเตือน: {dialog_state['message']} — ลองใหม่อีกครั้ง")
            else:
                print(f"  [เตือน] การกดค้นหา/รอโหลดหน้ามีปัญหา: {e}")
            continue

        if dialog_state["message"]:
            print(f"  [ไม่ผ่าน] เบราว์เซอร์แจ้งเตือน: {dialog_state['message']} — ลองใหม่อีกครั้ง")
            continue

        # เก็บร่องรอยไว้ช่วยวินิจฉัย: ถ้า CAPTCHA ผิดแบบไม่มี error แจ้งเลย
        # (เว็บไซต์ไม่แจ้งอะไร) ต้องดูว่าช่อง dfId ยังมีค่าอยู่ไหมหลังกดค้นหา
        # เพื่อเทียบว่าเป็นการ submit สำเร็จจริงหรือถูกปฏิเสธเงียบๆ
        try:
            dfid_after = page.input_value("input[name='dfId']")
            print(f"  [ตรวจสอบ] ค่า dfId หลังกดค้นหา: '{dfid_after}'")
        except Exception:
            pass

        return parse_results_table(page)

    return STATUS_ERROR, "พิมพ์ CAPTCHA ผิดครบจำนวนครั้งที่กำหนด หรือระบบไม่ยอมรับซ้ำๆ"


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default="CardID.xlsx", help="ไฟล์ Excel รายชื่อ/เลขบัตรประชาชน")
    parser.add_argument("--sheet", default="0769", help="ชื่อชีตในไฟล์ input")
    parser.add_argument("--id-col", type=int, default=0, help="index คอลัมน์เลขบัตรประชาชน (0=A)")
    parser.add_argument("--member-col", type=int, default=1, help="index คอลัมน์เลขทะเบียนสมาชิก (0=A)")
    parser.add_argument("--name-col", type=int, default=2, help="index คอลัมน์ชื่อ-นามสกุล (0=A)")
    parser.add_argument("--output", default="results.xlsx", help="ไฟล์ผลลัพธ์ (จะสร้าง/ต่อท้ายให้)")
    parser.add_argument("--limit", type=int, default=0, help="จำกัดจำนวนรายชื่อที่จะตรวจ (0 = ไม่จำกัด) ใช้สำหรับทดสอบ")
    parser.add_argument("--captcha-image", default="captcha.png", help="ไฟล์ที่จะบันทึกรูป CAPTCHA ล่าสุด")
    args = parser.parse_args()

    print(f"กำลังอ่านไฟล์ input: {args.input} (ชีต {args.sheet})")
    records = read_input_rows(args.input, args.sheet, args.id_col, args.member_col, args.name_col)
    print(f"พบรายชื่อทั้งหมด {len(records)} ราย")

    done_ids = load_done_ids(args.output)
    if done_ids:
        print(f"พบผลลัพธ์เดิมในไฟล์ output {len(done_ids)} ราย จะข้ามรายที่ตรวจไปแล้ว")
    remaining = [r for r in records if r["id"] not in done_ids]

    if args.limit:
        remaining = remaining[: args.limit]

    print(f"จะตรวจสอบทั้งหมด {len(remaining)} ราย ในรอบนี้\n")
    if not remaining:
        print("ไม่มีรายชื่อที่ต้องตรวจสอบเพิ่ม")
        return

    wb, ws = open_or_create_output(args.output)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        dialog_state = setup_dialog_capture(page)

        print("=" * 60)
        print("ในหน้าต่างเบราว์เซอร์ที่เปิดขึ้นมา กรุณาพิมพ์ URL แล้วนำทางไปด้วยตนเอง")
        print("เหมือนที่คุณทำเป็นปกติ จนกว่าจะเข้าสู่ระบบและเปิดหน้า")
        print("'สอบถามบุคคลล้มละลาย' (WEB3Q010) ได้สำเร็จ 1 ครั้ง")
        print("เสร็จแล้วกลับมาที่ terminal นี้แล้วกด Enter เพื่อเริ่มตรวจสอบอัตโนมัติ")
        print("=" * 60)
        input()

        counts = {STATUS_CLEAR: 0, STATUS_FOUND: 0, STATUS_ERROR: 0}
        total = len(remaining)
        for idx, record in enumerate(remaining, start=1):
            print(f"\n[{idx}/{total}] ===================================")
            try:
                status, detail = process_one(page, record, args.captcha_image, dialog_state)
            except KeyboardInterrupt:
                print("\nหยุดโดยผู้ใช้ — ผลลัพธ์ที่ตรวจไปแล้วถูกบันทึกไว้แล้ว")
                break
            except Exception as e:
                status, detail = STATUS_ERROR, f"เกิดข้อผิดพลาด: {e}"

            append_result(wb, ws, args.output, record, status, detail)
            counts[status] = counts.get(status, 0) + 1
            print(f"  ผลลัพธ์: {status} {('- ' + detail) if detail else ''}")

        browser.close()

        print("\n" + "=" * 60)
        print("สรุปผลรอบนี้")
        for k, v in counts.items():
            print(f"  {k}: {v} ราย")
        print(f"บันทึกผลลัพธ์ทั้งหมดไว้ที่: {args.output}")


if __name__ == "__main__":
    main()
